"""
excel/qa_logger.py — Client Q&A Report Generator

PURPOSE:
  When the user says "export to excel", this creates a professional
  2-column Excel report from all questions asked in the chat session:

      Column A  →  Question  (what the client asked)
      Column B  →  Answer    (what the bot answered from the data)

  This is the PRIMARY deliverable — a clean, client-ready report.

WORKFLOW:
  1. User asks questions to the bot in chat
  2. User types: "export to excel"
  3. This module reads all Q&A pairs from the session memory
  4. Creates a formatted Excel file saved to output/qa_report.xlsx
  5. Download button appears in the sidebar

NOTE:
  This replaces the fixed-template approach. The data source (resume/PDF)
  can be swapped — the pipeline stays the same, only the PDF changes.
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ─── Color Palette ──────────────────────────────────────────────────────────
COLOR_BRAND_DARK   = "1E3A5F"   # Deep navy — title bar
COLOR_BRAND_MAIN   = "2563EB"   # Blue — question header
COLOR_BRAND_ACCENT = "059669"   # Green — answer header
COLOR_Q_ROW_ODD    = "EFF6FF"   # Light blue — question cells (odd rows)
COLOR_Q_ROW_EVEN   = "DBEAFE"   # Slightly darker blue (even rows)
COLOR_A_ROW_ODD    = "F0FDF4"   # Light green — answer cells (odd rows)
COLOR_A_ROW_EVEN   = "DCFCE7"   # Slightly darker green (even rows)
COLOR_BORDER       = "CBD5E1"   # Subtle grey border


def _make_border(color: str = COLOR_BORDER) -> Border:
    side = Side(style="thin", color=color)
    return Border(top=side, bottom=side, left=side, right=side)


def _make_thick_bottom(color: str = COLOR_BORDER) -> Border:
    return Border(
        top=Side(style="thin", color=color),
        bottom=Side(style="medium", color="94A3B8"),
        left=Side(style="thin", color=color),
        right=Side(style="thin", color=color),
    )


def export_session_to_excel(qa_pairs: list[tuple[str, str]], filepath: str) -> int:
    """
    Create a professional 2-column Q&A report Excel file.

    Layout:
        Row 1       : Merged title bar  (Company/project name + date)
        Row 2       : Column headers    (Question | Answer)
        Row 3+      : Data rows         (one Q&A pair per row)

    Args:
        qa_pairs  : List of (question, answer) tuples from the chat session
        filepath  : Save path (from config.QA_LOG_PATH)

    Returns:
        Number of Q&A rows written
    """
    if not qa_pairs:
        return 0

    os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Q&A Report"

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 52   # Question column
    ws.column_dimensions["B"].width = 78   # Answer column

    # ── ROW 1: Title bar ─────────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    date_str = datetime.now().strftime("%d %B %Y")
    title_cell.value = f"Client Q&A Report  |  Generated on {date_str}"
    title_cell.font = Font(
        name="Calibri", bold=True, size=14, color="FFFFFF"
    )
    title_cell.fill = PatternFill(
        start_color=COLOR_BRAND_DARK,
        end_color=COLOR_BRAND_DARK,
        fill_type="solid"
    )
    title_cell.alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.row_dimensions[1].height = 36

    # ── ROW 2: Column headers ─────────────────────────────────────────────────
    header_data = [
        ("A", "Question", COLOR_BRAND_MAIN),
        ("B", "Answer",   COLOR_BRAND_ACCENT),
    ]
    for col_letter, label, color in header_data:
        cell = ws[f"{col_letter}2"]
        cell.value = label
        cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        cell.border = _make_thick_bottom()
    ws.row_dimensions[2].height = 30

    # ── ROW 3+: Data rows ─────────────────────────────────────────────────────
    wrap_align = Alignment(vertical="top", wrap_text=True, horizontal="left")

    for idx, (question, answer) in enumerate(qa_pairs):
        row = idx + 3          # Data starts at row 3
        is_odd = (idx % 2 == 0)

        q_color = COLOR_Q_ROW_ODD  if is_odd else COLOR_Q_ROW_EVEN
        a_color = COLOR_A_ROW_ODD  if is_odd else COLOR_A_ROW_EVEN

        # Column A — Question
        q_cell = ws.cell(row=row, column=1, value=question)
        q_cell.font      = Font(name="Calibri", size=11, color="1E3A5F", bold=True)
        q_cell.fill      = PatternFill(start_color=q_color, end_color=q_color, fill_type="solid")
        q_cell.alignment = wrap_align
        q_cell.border    = _make_border()

        # Column B — Answer
        a_cell = ws.cell(row=row, column=2, value=answer)
        a_cell.font      = Font(name="Calibri", size=11, color="1A3C2B")
        a_cell.fill      = PatternFill(start_color=a_color, end_color=a_color, fill_type="solid")
        a_cell.alignment = wrap_align
        a_cell.border    = _make_border()

        # Row height — scale with content length
        max_chars = max(len(question), len(answer))
        row_height = min(max(40, max_chars // 3), 200)
        ws.row_dimensions[row].height = row_height

    # ── Freeze header rows so they stay visible when scrolling ───────────────
    ws.freeze_panes = "A3"   # Freeze title + header rows

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:B{len(qa_pairs) + 2}"

    wb.save(filepath)
    print(f"[OK] Client Q&A report saved: {filepath} ({len(qa_pairs)} rows)")
    return len(qa_pairs)
