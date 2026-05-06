"""
excel/writer.py — Excel Answer Writer

WHAT IT DOES:
  Takes the {field_name: answer} dictionary produced by mapper.py,
  opens the Excel template, finds each field in Column A, and writes
  the corresponding answer into Column B.

  Saves the result to output/filled_resume.xlsx — the template is
  NEVER modified (we always write to a separate output file).

WHY openpyxl (not xlwings or pandas):
  - No Microsoft Excel installation required — pure Python
  - Preserves all template formatting: colors, borders, fonts
  - Works on Windows, Mac, and Linux identically
  - Lightweight and fast

CELL FORMATTING APPLIED:
  Filled cells get a light green background so it's visually clear
  which cells were auto-filled vs. left empty.
"""

import os
import openpyxl
from openpyxl.styles import PatternFill, Font
from copy import copy
from pathlib import Path


# Light green fill for auto-filled cells
FILLED_CELL_COLOR = "D6F4D6"
NOT_FOUND_COLOR   = "FFF3CD"   # Light yellow for "not found" cells

FILLED_FILL  = PatternFill(start_color=FILLED_CELL_COLOR, end_color=FILLED_CELL_COLOR, fill_type="solid")
MISSING_FILL = PatternFill(start_color=NOT_FOUND_COLOR,  end_color=NOT_FOUND_COLOR,  fill_type="solid")


def fill_excel(
    field_answer_map: dict,
    template_path: str,
    output_path: str,
) -> str:
    """
    Write answers into the Excel template and save as output file.

    Args:
        field_answer_map : {field_name: answer_or_None} from mapper.py
        template_path    : Path to the Excel template (read-only source)
        output_path      : Path where the filled Excel will be saved

    Returns:
        output_path (for confirmation)
    """
    if not Path(template_path).exists():
        raise FileNotFoundError(f"Excel template not found: {template_path}")

    # Load the template (preserves all formatting)
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    filled_count = 0

    for row in ws.iter_rows(min_row=1):
        cell_a = row[0]   # Field name cell (Column A)
        cell_b = row[1]   # Answer cell (Column B)

        if not cell_a.value:
            continue

        field_name = str(cell_a.value).strip()

        # Skip section headers (they start with "SECTION:")
        if field_name.upper().startswith("SECTION:"):
            continue

        # Write the answer if we have one
        if field_name in field_answer_map:
            answer = field_answer_map[field_name]

            if answer:
                cell_b.value = answer
                cell_b.fill  = FILLED_FILL
                cell_b.font  = Font(bold=False, color="1A5C1A")  # Dark green text
                filled_count += 1
            else:
                cell_b.value = "Not found in resume"
                cell_b.fill  = MISSING_FILL

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb.save(output_path)
    print(f"[OK] Excel saved -> {output_path}  ({filled_count} cells filled)")
    return output_path
