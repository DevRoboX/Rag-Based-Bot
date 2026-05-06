"""
setup.py — One-Time Setup Script

RUN THIS ONCE before launching the chatbot.

WHAT IT DOES:
  1. Creates the Excel template with all resume fields
  2. Loads your resume file from ./data/
  3. Splits it into chunks
  4. Embeds chunks using BAAI/bge-base-en-v1.5
  5. Persists them in ChromaDB at ./chroma_db/

USAGE:
    python setup.py                    # Full setup (template + ingest)
    python setup.py --ingest-only      # Only re-ingest resume (skip template)
    python setup.py --template-only    # Only create Excel template
"""

import sys
import os
import argparse


def create_excel_template():
    """Create the Excel template with predefined resume fields."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from config import EXCEL_TEMPLATE_PATH

    os.makedirs(os.path.dirname(EXCEL_TEMPLATE_PATH), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resume Data"

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    # Style definitions
    header_fill   = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    section_fill  = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    header_font   = Font(bold=True, color="FFFFFF", size=12)
    section_font  = Font(bold=True, color="3730A3", size=11)
    label_font    = Font(bold=False, color="111827", size=10)
    thin_border   = Border(
        bottom=Side(style="thin", color="E5E7EB"),
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
    )

    # ── Column Headers (Row 1) ──
    ws["A1"] = "Field"
    ws["B1"] = "Answer (Auto-filled by RAG Bot)"
    for cell in [ws["A1"], ws["B1"]]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Resume Sections & Fields ──
    sections = [
        ("SECTION: Contact Information", [
            "Full Name",
            "Email Address",
            "Phone Number",
            "Location",
            "LinkedIn Profile",
            "GitHub Profile",
            "Portfolio / Website",
        ]),
        ("SECTION: Professional Summary", [
            "Professional Summary",
            "Career Objective",
        ]),
        ("SECTION: Skills", [
            "Technical Skills",
            "Programming Languages",
            "Tools & Frameworks",
            "Soft Skills",
        ]),
        ("SECTION: Work Experience", [
            "Years of Experience",
            "Current / Last Company",
            "Current / Last Role",
            "Current / Last Duration",
            "Key Responsibilities",
            "Achievements at Work",
            "Previous Company",
            "Previous Role",
            "Previous Duration",
        ]),
        ("SECTION: Education", [
            "Highest Degree",
            "University / College",
            "Field of Study / Major",
            "Graduation Year",
            "CGPA / GPA / Percentage",
        ]),
        ("SECTION: Projects", [
            "Key Projects",
            "Project Technologies",
        ]),
        ("SECTION: Certifications", [
            "Certifications",
            "Online Courses",
        ]),
        ("SECTION: Achievements", [
            "Awards & Recognition",
            "Publications",
        ]),
        ("SECTION: Additional Info", [
            "Languages Known",
            "Hobbies / Interests",
            "References",
        ]),
    ]

    current_row = 2
    for section_label, fields in sections:
        # Section header row
        ws.cell(row=current_row, column=1, value=section_label)
        ws.cell(row=current_row, column=2, value="")
        for col in [1, 2]:
            cell = ws.cell(row=current_row, column=col)
            cell.fill      = section_fill
            cell.font      = section_font
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[current_row].height = 22
        ws.merge_cells(f"A{current_row}:B{current_row}")
        current_row += 1

        # Field rows
        for field in fields:
            ws.cell(row=current_row, column=1, value=field).font   = label_font
            ws.cell(row=current_row, column=2, value="").font      = label_font
            for col in [1, 2]:
                ws.cell(row=current_row, column=col).border    = thin_border
                ws.cell(row=current_row, column=col).alignment = Alignment(vertical="center", wrap_text=True)
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1  # Blank row between sections

    wb.save(EXCEL_TEMPLATE_PATH)
    print(f"[OK] Excel template created -> {EXCEL_TEMPLATE_PATH}")


def ingest_resume():
    """Run the full ingestion pipeline: load → chunk → embed → store."""
    from config import DATA_DIR, CHUNK_SIZE, CHUNK_OVERLAP
    from ingestion.loader import load_all_from_dir
    from ingestion.chunker import chunk_documents
    from ingestion.embedder import build_vector_store

    print("\n" + "="*50)
    print("RESUME INGESTION PIPELINE")
    print("="*50)

    # Step 1: Load resume(s)
    print("\n[Step 1/3] Loading resume file(s)...")
    documents = load_all_from_dir(DATA_DIR)

    # Step 2: Chunk documents
    print("\n[Step 2/3] Chunking documents...")
    chunks = chunk_documents(documents, chunk_size=CHUNK_SIZE, chunk_overlap=CHUNK_OVERLAP)

    # Step 3: Embed and store in ChromaDB
    print("\n[Step 3/3] Embedding and storing in ChromaDB...")
    build_vector_store(chunks)

    print("\n" + "="*50)
    print("INGESTION COMPLETE!")
    print("   Run the chatbot with:  streamlit run ui/app.py")
    print("="*50 + "\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="RAG Resume Bot Setup")
    parser.add_argument("--ingest-only",   action="store_true", help="Only run ingestion")
    parser.add_argument("--template-only", action="store_true", help="Only create Excel template")
    args = parser.parse_args()

    if args.ingest_only:
        ingest_resume()
    elif args.template_only:
        create_excel_template()
    else:
        print("Running full setup...\n")
        create_excel_template()
        ingest_resume()
